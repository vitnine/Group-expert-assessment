# _*_ coding: utf-8 _*_
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from os.path import join, abspath

from models import get_matrix, get_head, get_wall

data_path = join('.', "matrix.xlsx")
data_path = abspath(data_path)

wb = load_workbook(filename=data_path, data_only=True, read_only=True)

work_sheets_names = wb.sheetnames

work_sheets_data = None

print(work_sheets_names)
user_sheet = input("\nChoose your sheet: ")

while user_sheet not in work_sheets_names:
    print("Unknown sheet, try again")
    user_sheet = input("Choose your sheet")

print(f"{user_sheet} was chosen successfully")
ws = wb[user_sheet]

wall = get_wall(ws)
head = get_head(ws)
matrix = get_matrix(ws)

print(head)
print(wall)
print(matrix)


