from openpyxl import load_workbook, Workbook
from utils.auxiliary import CorrectPath

work_book = load_workbook(filename=CorrectPath.matrix_path, data_only=True, read_only=True)
work_sheets_names = work_book.sheetnames
work_sheets_data = None
